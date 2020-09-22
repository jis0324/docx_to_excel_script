import openpyxl
import re
import os
import comtypes.client
import PyPDF2 

base_dir = os.path.dirname(os.path.abspath(__file__))
input_docx_path = base_dir + '/FISH Assessment.docx'
output_xlsx_path = base_dir + '/result.xlsx'

# convert docx to pdf
def convert_docx_to_pdf():
  out_pdf_path = base_dir + '/output.pdf'
  wdFormatPDF = 17

  word = comtypes.client.CreateObject('Word.Application')
  doc = word.Documents.Open(input_docx_path)
  doc.SaveAs(out_pdf_path, FileFormat=wdFormatPDF)
  doc.Close()
  word.Quit()

# extract the lesson data such as task, 
def get_lesson_data(content):
  # define return dict
  return_dict = dict()
  return_dict['title'] = ''
  return_dict['task'] = ''
  return_dict['prerequisites'] = ''
  return_dict['concept'] = ''
  return_dict['behavioral_objective'] = ''
  return_dict['materials'] = ''
  return_dict['task_analysis'] = ''

  task_keyword_pattern = re.compile(r'Task\s*\:')
  if task_keyword_pattern.search(content):
    task_keyword_start_pos = task_keyword_pattern.search(content).start()
    task_keyword_end_pos = task_keyword_pattern.search(content).end()
    title = content[:task_keyword_start_pos].strip()
    return_dict['title'] = re.sub(r'\s+', ' ', title)
    content = content[task_keyword_end_pos:].strip()
    
    prerequisites_keyword_pattern = re.compile(r'Prerequisites\s*\:')
    if prerequisites_keyword_pattern.search(content):
      prerequisites_keyword_start_pos = prerequisites_keyword_pattern.search(content).start()
      prerequisites_keyword_end_pos = prerequisites_keyword_pattern.search(content).end()
      task = content[:prerequisites_keyword_start_pos].strip()
      return_dict['task'] = re.sub(r'\s+', ' ', task)
      content = content[prerequisites_keyword_end_pos:].strip()
    
    concept_keyword_pattern = re.compile(r'Concept\s*\:')
    if concept_keyword_pattern.search(content):
      concept_keyword_start_pos = concept_keyword_pattern.search(content).start()
      concept_keyword_end_pos = concept_keyword_pattern.search(content).end()
      prerequisites = content[:concept_keyword_start_pos].strip()
      return_dict['prerequisites'] = re.sub(r'\s+', ' ', prerequisites)
      content = content[concept_keyword_end_pos:].strip()
    
    behavioral_keyword_pattern = re.compile(r'Behavioral\s+Objective\s*\:')
    if behavioral_keyword_pattern.search(content):
      behavioral_keyword_start_pos = behavioral_keyword_pattern.search(content).start()
      behavioral_keyword_end_pos = behavioral_keyword_pattern.search(content).end()
      concept = content[:behavioral_keyword_start_pos].strip()
      return_dict['concept'] = re.sub(r'\s+', ' ', concept)
      content = content[behavioral_keyword_end_pos:].strip()

    materials_keyword_pattern = re.compile(r'Materials\s*\:')
    if materials_keyword_pattern.search(content):
      materials_keyword_start_pos = materials_keyword_pattern.search(content).start()
      materials_keyword_end_pos = materials_keyword_pattern.search(content).end()
      behavioral_objective = content[:materials_keyword_start_pos].strip()
      return_dict['behavioral_objective'] = re.sub(r'\s+', ' ', behavioral_objective)
      content = content[materials_keyword_end_pos:].strip()

    analysis_keyword_pattern = re.compile(r'Task\s+Ana\s*lysis\s*\:')
    if analysis_keyword_pattern.search(content):
      analysis_keyword_start_pos = analysis_keyword_pattern.search(content).start()
      analysis_keyword_end_pos = analysis_keyword_pattern.search(content).end()
      materials = content[:analysis_keyword_start_pos].strip()
      return_dict['materials'] = re.sub(r'\s+', ' ', materials)
      content = content[analysis_keyword_end_pos:].strip()

    newline_pos_list = list()
    content = re.sub(r'\s+', ' ', content).strip()
    row_num_pattern = re.compile(r'\.\s+\d+\.\s')
    for row in row_num_pattern.finditer(content):
      newline_pos_list.append(row.start())
    
    temp_list = list()
    for index in range(len(newline_pos_list) + 1):
      if index == 0:
        temp_list.append(content[:newline_pos_list[index]])
      else:
        if len(newline_pos_list) > index:
          temp_list.append(content[newline_pos_list[index-1]:newline_pos_list[index]])
        else:
          temp_list.append(content[newline_pos_list[index-1]:len(content)-10])

    return_dict['task_analysis'] = "\n".join(temp_list)

    return return_dict
  else:
    return None

# create result.xlsx
def create_xlsx():
  # Create a Workbook
  wb = openpyxl.Workbook()
  ws =  wb.active
  ws.title = "Lessons"

  # write fieldnames
  ws.cell(row=1, column=1).value = "Lesson"
  ws.cell(row=1, column=2).value = "Task"
  ws.cell(row=1, column=3).value = "Prerequisites"
  ws.cell(row=1, column=4).value = "Concept"
  ws.cell(row=1, column=5).value = "Behavioral Objective"
  ws.cell(row=1, column=6).value = "Materials"
  ws.cell(row=1, column=7).value = "Task Analysis"
  wb.save(output_xlsx_path)

# insert lesson to xlsx
def insert_to_xlsx(row_dict):
  file_exist = os.path.isfile(output_xlsx_path)
  if not file_exist:
    # create new result.xlsx
    create_xlsx()

  # excel load
  wb_obj = openpyxl.load_workbook(output_xlsx_path)

  # from the active attribute 
  sheet_obj = wb_obj.active

  # get max column count
  max_row = sheet_obj.max_row
  row_index = max_row + 1
  
  sheet_obj.cell(row=row_index, column=1).value = row_dict["title"]
  sheet_obj.cell(row=row_index, column=2).value = row_dict["task"]
  sheet_obj.cell(row=row_index, column=3).value = row_dict["prerequisites"]
  sheet_obj.cell(row=row_index, column=4).value = row_dict["concept"]
  sheet_obj.cell(row=row_index, column=5).value = row_dict["behavioral_objective"]
  sheet_obj.cell(row=row_index, column=6).value = row_dict["materials"]
  sheet_obj.cell(row=row_index, column=7).value = row_dict["task_analysis"]

  wb_obj.save(output_xlsx_path)

# formatting result xlsx
def format_xlsx():
  # excel load
  wb_obj = openpyxl.load_workbook(output_xlsx_path)
  # from the active attribute 
  sheet_obj = wb_obj.active

  for row_index in range(sheet_obj.max_row):
    for col in sheet_obj.columns:
      sheet_obj["{}{}".format(col[0].column, row_index + 1)].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True) 
    
    if row_index == 0:
      continue
    sheet_obj.row_dimensions[row_index + 1].height = 100

  sheet_obj.column_dimensions['A'].width = 10
  sheet_obj.column_dimensions['B'].width = 20
  sheet_obj.column_dimensions['C'].width = 20
  sheet_obj.column_dimensions['D'].width = 40
  sheet_obj.column_dimensions['E'].width = 40
  sheet_obj.column_dimensions['F'].width = 40
  sheet_obj.column_dimensions['G'].width = 50

  wb_obj.save(output_xlsx_path)

# main
def main():
  if input_docx_path.rsplit('.', 1)[1] == "docx":
    print("Converting docx to pdf.....")
    convert_docx_to_pdf()
  elif input_docx_path.rsplit('.', 1)[1] != "pdf":
    print("----- Unkown File Type, Try again. -----")
    return

  # open pdf file
  pdfFileObj = open(base_dir + '/output.pdf', 'rb')

  # creating a pdf reader object
  pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
    
  # printing number of pages in pdf file 
  for page in range(pdfReader.numPages):
    print("Converting {} page in Total {}".format(page + 1, pdfReader.numPages))

    # creating a page object 
    pageObj = pdfReader.getPage(page)
    
    # extracting text from page 
    page_text = pageObj.extractText()

    if page_text.strip():
      # extract page data
      lesson_dict = get_lesson_data(page_text)

      if lesson_dict:
        insert_to_xlsx(lesson_dict)
  
  # formating result xlsx
  format_xlsx()

  # closing the pdf file object 
  pdfFileObj.close()

  # delete pdf file
  os.remove(base_dir + '/output.pdf')

if __name__ == "__main__":
  # delete output xlsx
  if os.path.isfile(output_xlsx_path):
    os.remove(output_xlsx_path)

  main()
